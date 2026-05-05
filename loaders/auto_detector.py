"""Automatic loader detection."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from ..core.exceptions import UnsupportedFormatError
from ..utils.validators import ensure_list_of_dicts
from .csv_loader import CSVLoader
from .json_loader import JSONLoader


class AutoDetector:
    """Detect supported file and in-memory input types."""

    def __init__(self) -> None:
        self.json_loader = JSONLoader()
        self.csv_loader = CSVLoader()

    def load(self, source: Any) -> list[dict[str, Any]]:
        if isinstance(source, (list, tuple, dict)) or hasattr(source, "to_dict"):
            return ensure_list_of_dicts(source)

        if isinstance(source, (str, Path)):
            path = Path(source)
            suffix = path.suffix.lower()
            if suffix in {".json", ".jsonl"}:
                return self.json_loader.load(path)
            if suffix == ".csv":
                return self.csv_loader.load(path)
            if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                return self._load_excel(path)
            raise UnsupportedFormatError(
                f"Unsupported file format: {suffix or '[no extension]'}"
            )

        return ensure_list_of_dicts(source)

    def _load_excel(self, path: Path) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UnsupportedFormatError(
                "Excel support requires openpyxl to be installed."
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(header) if header is not None else f"column_{index}" for index, header in enumerate(rows[0])]
        data: list[dict[str, Any]] = []
        for row in rows[1:]:
            data.append({headers[index]: value for index, value in enumerate(row)})
        return data
